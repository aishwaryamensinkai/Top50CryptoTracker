import requests
import pandas as pd
import time
from openpyxl import load_workbook
import openpyxl

# Replace 'your_api_key' with your actual CoinMarketCap API key
api_key = 'your-api-key'

# Function to fetch live data from CoinMarketCap API
def fetch_crypto_data():
    url = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest'
    
    # Headers including the API key
    headers = {
        'Accepts': 'application/json',
        'X-CMC_PRO_API_KEY': api_key,
    }
    
    # Parameters to get the top 50 cryptocurrencies by market cap
    parameters = {
        'start': '1',
        'limit': '50',
        'convert': 'USD'
    }
    
    # Making the API request
    response = requests.get(url, headers=headers, params=parameters)
    
    # Check if the request was successful
    if response.status_code == 200:
        data = response.json()['data']
    else:
        print(f"Error fetching data: {response.status_code}")
        return pd.DataFrame()
    
    # Extract relevant fields from the data
    crypto_data = []
    for crypto in data:
        crypto_data.append({
            'Name': crypto['name'],
            'Symbol': crypto['symbol'],
            'Current Price (USD)': crypto['quote']['USD']['price'],
            'Market Cap': crypto['quote']['USD']['market_cap'],
            '24h Volume': crypto['quote']['USD']['volume_24h'],
            '24h Change (%)': crypto['quote']['USD']['percent_change_24h']
        })
    
    return pd.DataFrame(crypto_data)

# Function to perform analysis on the crypto data
def analyze_crypto_data(df):
    # Identify the top 5 cryptocurrencies by market cap
    top_5_by_market_cap = df.nlargest(5, 'Market Cap')

    # Calculate the average price of the top 50 cryptocurrencies
    average_price = df['Current Price (USD)'].mean()

    # Analyze the highest and lowest 24-hour percentage price change
    highest_change = df.loc[df['24h Change (%)'].idxmax()]
    lowest_change = df.loc[df['24h Change (%)'].idxmin()]

    return top_5_by_market_cap, average_price, highest_change, lowest_change

# Function to update the Excel sheet with live data
def update_excel_sheet(df):
    file_name = "Live_Crypto_Data.xlsx"
    
    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    
    # Check if sheet exists, if not create one
    if 'Crypto Data' not in workbook.sheetnames:
        sheet = workbook.create_sheet(title='Crypto Data')
    else:
        sheet = workbook['Crypto Data']
    
    # Clear existing data
    sheet.delete_rows(1, sheet.max_row)

    # Write headers
    headers = list(df.columns)
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)
    
    # Write data rows
    for row_num, row in enumerate(df.values, start=2):
        for col_num, value in enumerate(row, start=1):
            sheet.cell(row=row_num, column=col_num, value=value)

    workbook.save(file_name)
    print("Excel sheet updated")

# Main function to run the script
def main():
    while True:
        print("Fetching live data...")
        df = fetch_crypto_data()
        if not df.empty:
            update_excel_sheet(df)
        
            # Perform data analysis
            top_5_by_market_cap, average_price, highest_change, lowest_change = analyze_crypto_data(df)
            
            # Display analysis results in the console
            print("\nTop 5 Cryptocurrencies by Market Cap:\n", top_5_by_market_cap)
            print("\nAverage Price of Top 50 Cryptocurrencies: $", round(average_price, 2))
            print("\nHighest 24h Change:\n", highest_change)
            print("\nLowest 24h Change:\n", lowest_change)
        
        # Wait 5 minutes before updating again
        time.sleep(300)

if __name__ == "__main__":
    main()
